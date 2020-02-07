# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Alignment

'''
aggregate report:
金額意義上名為價格,價值 ; 收費金額合計為總共收到的實質金錢
如果用兌換券, 金額(價值)一樣會寫上每堂課的價格, 但實收金額為0;

Q: 換課:原報名活動=.... 跟換課作廢,已換至xxxx 是否一樣意思 ?
A: 換課作廢,已換至... 的是被換掉不要的課(before), 換課: 原報名活動.... 為換成的課(after). 被換掉的課報名狀態為已換課, 換成的課狀態一樣為完成繳費, 註記是被換過來的課(source + dest note)

'''

g_rows = None
g_columns = None
g_workbook = None
g_worksheet  = None

g_student_data = {}
g_transaction_state = {
    u'完成繳費':'complete', #備註若為使用兌換券系統實收會是0
    u'已換課':'change_class',
    u'取消':'cancel', # 金額不列入實收
    u'退費-因事':'refund_issue',
    u'退費-因病':'refund_sick',
    u'退費-中心因素':'refund_center',
}

g_transaction_desc = {}

def init():
    global g_rows, g_columns, g_workbook, g_worksheet, g_transaction_desc
    g_workbook = openpyxl.load_workbook('export.xlsx')
    g_worksheet = g_workbook.get_sheet_by_name('Sheet1')
    g_rows = list(g_worksheet.rows) # list type
    g_columns = list(g_worksheet.columns)

    for k,v in g_transaction_state.items():
        g_transaction_desc[v] = k

    print "Total rows={}, columns={}".format(len(g_rows), len(g_columns))

def get_transaction_desc(trans_state):
    pass

def get_cell(row, column, sheet):
    return sheet.cell(row=row,column=column).value

def get_row(row):
    return g_rows[row] # type = tuple of Cell object

def get_column(column):
    return g_columns[column]

def set_cell(row, column, sheet, value):
    sheet.cell(row=row,column=column).value = value

def parse_export():
    for idx in range(len(g_rows)):
        if idx == 0: # header
            continue
        r = g_rows[idx]
        class_name_tokens = r[0].value.split('-')
        class_name = class_name_tokens[0]
        #if len(class_name_tokens) > 1:
        #    print "class parent name-", class_name, "sub name:", class_name_tokens[1]

        student_name = r[1].value
        activity_date = r[2].value
        transaction_state = r[3].value
        if r[3].value not in g_transaction_state:
            print u"Error!{}".format(r[3].value)
            raise Exception(u'Error! Cannot find transaction state {}'.format(r[3].value))
        else:
            transaction_state = g_transaction_state[r[3].value]
        price = r[4].value
        if price == None:
            print u"警告!!!!!!!!!!!!!!:學生:{} 課程:{} 無金額資訊!!".format(student_name, r[0].value)
            price = 0
        actual_pay_amount = r[5].value
        note = r[6].value if r[6].value != None else ""
        change_class_note = r[7].value
        if student_name not in g_student_data:
            g_student_data[student_name] = {
                'classes':{},
            }

        s_info = g_student_data[student_name]

        if class_name not in s_info['classes']:
            s_info['classes'][class_name] = []

        data = {'date':activity_date,
                'price':price, 
                'pay_amount': actual_pay_amount,
                'transaction_state':transaction_state,
                'note':note,
                'change_class_note': change_class_note,
                'whole_class_name':r[0].value,
        }
        s_info['classes'][class_name].append(data)

def aggregate_export():
    merge_info = []
    workbook = openpyxl.Workbook()
    sheet=workbook.active
    sheet['A1']=u'請自行輸入主題' # column 1
    sheet['A2']=u'編號' # column 1
    sheet['B2']=u'姓名' # column 2
    sheet['C2']=u'課程名稱' # column 3
    sheet['D2']=u'課程日期' # column 4
    sheet['E2']=u'堂數' # column 5
    sheet['F2']=u'金額' # column 6
    sheet['G2']=u'收費金額合計' # column 7
    sheet['H2']=u'退費金額' # column 8
    sheet['I2']=u'實收合計' # column 9
    sheet['J2']=u'備註(退費/退費說明)' # column 10
    sheet['K2']=u'收據編號' # column 11

    student_no = 1
    current_row = 3
    for student_name, data in g_student_data.items():
        print u"Student={}".format(student_name)
        
        sheet.cell(row=current_row, column=1).value = student_no
        sheet.cell(row=current_row, column=2).value = student_name
        
        class_count = len(data['classes'].values())
        idx = 0
        # class_data is an array of class data dictionary
        for class_name, class_data in data['classes'].items():
            print "class name=", class_name
            sheet.cell(row=current_row+idx, column=1).value = student_no ; sheet.cell(row=current_row+idx, column=1).alignment = Alignment(vertical='center')
            sheet.cell(row=current_row+idx, column=2).value = student_name ; sheet.cell(row=current_row+idx, column=2).alignment = Alignment(vertical='center')
            sheet.cell(row=current_row+idx, column=3).value = class_name
            class_dates = []
            notes = []
            total_price = 0 # 兌換券金額算在內
            total_pay_amount = 0
            total_refund = 0
            actual_pay_amount = 0
            for i in class_data:
                class_dates.append(i['date'])
                total_price+=int(i['price'])
                total_pay_amount+=i['pay_amount']
                if i['note'] != "":
                    notes.append(u"[{}:{}:{}]".format(i['date'], i['whole_class_name'], i['note']))
            class_dates.sort()
            class_date_str = u"、".join(i for i in class_dates)
            notes_str = u'、\n'.join(i for i in notes)

            sheet.cell(row=current_row+idx, column=4).value = class_date_str
            sheet.cell(row=current_row+idx, column=5).value = len(class_dates)
            sheet.cell(row=current_row+idx, column=6).value = total_price
            sheet.cell(row=current_row+idx, column=7).value = total_pay_amount
            sheet.cell(row=current_row+idx, column=8).value = total_refund
            sheet.cell(row=current_row+idx, column=9).value = actual_pay_amount ; sheet.cell(row=current_row+idx, column=9).alignment = Alignment(vertical='center')
            sheet.cell(row=current_row+idx, column=10).value = notes_str
            idx +=1

        if class_count > 1:
            #print "class count > 1,", student_name, class_count
            merge_info.append((current_row, current_row+class_count-1, 1, 1)) # start_row, end_row, start_column, end_column
            merge_info.append((current_row, current_row+class_count-1, 2, 2)) 
            merge_info.append((current_row, current_row+class_count-1, 9, 9))
            merge_info.append((current_row, current_row+class_count-1, 11, 11))

        current_row += class_count
        student_no+=1
    
    #for i in merge_info[:2]:
    for i in merge_info:
        #sheet.merge_cells(start_row=current_row, end_row=current_row + class_count, start_column=1, end_column=1) # column: A to A 
        sheet.merge_cells(start_row=i[0], end_row=i[1], start_column=i[2], end_column=i[3])

    workbook.save('new.xlsx')


def print_student_data():
    for student_name, data in g_student_data.items():
        print u"Student={}".format(student_name)
        for class_name, class_data in data['classes'].items():
            for i in class_data:
                print u"Class={}, Date={}, Amount={},State={}".format(class_name, i['date'], i['pay_amount'], g_transaction_desc[i['transaction_state']])


def merge_cell_test():
    workbook = openpyxl.Workbook()
    sheet=workbook.active
    sheet['A1']='A1'
    sheet['A2']='A2'
    sheet['B1']='B1'
    sheet['B2']='B2'

    x1 = list(sheet.rows)[0]
    x2 = list(sheet.rows)[1]

    #print x1[0].value
    #print x1[1].value
    #print x2[0].value
    #print x2[1].value

    # When you merge cells all cells but the top-left one are removed from the worksheet.

    sheet.merge_cells(start_row=1, end_row=2, start_column=1, end_column=1)
    # equal: so merge cell is just a view manipulation problem
    #print x1[0].value
    #print x1[1].value
    #print x2[0].value
    #print x2[1].value
    workbook.save('new.xlsx')

if __name__ == "__main__":
    init()
    parse_export()
    #print_student_data()
    aggregate_export()
    #merge_cell_test()